#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu May 28 09:19:15 2020

@author: YOUSSEF NAAINIA
"""


from openpyxl import *
from tkinter import *
  

wb  = load_workbook('1ER.xlsx')
wb2 = load_workbook('2EME.xlsx')
wb3 = load_workbook('LP.xlsx')

sheet = wb.active
sheet2 = wb2.active
sheet3 = wb3.active

  
  
def excel(): 
      
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
  
    sheet.cell(row=1, column=1).value = "Nom"
    sheet.cell(row=1, column=2).value = "Prénom"
    sheet.cell(row=1, column=3).value = "CIN"
    sheet.cell(row=1, column=4).value = "CNE"
    sheet.cell(row=1, column=5).value = "Email"
    sheet.cell(row=1, column=6).value = "Téléphone"
    sheet.cell(row=1, column=7).value = "Classe"
    sheet.cell(row=1, column=8).value = "Filière"

    sheet2.column_dimensions['A'].width = 30
    sheet2.column_dimensions['B'].width = 10
    sheet2.column_dimensions['C'].width = 10
    sheet2.column_dimensions['D'].width = 20
    sheet2.column_dimensions['E'].width = 20
    sheet2.column_dimensions['F'].width = 40
    sheet2.column_dimensions['G'].width = 50

    sheet2.cell(row=1, column=1).value = "Nom"
    sheet2.cell(row=1, column=2).value = "Prénom"
    sheet2.cell(row=1, column=3).value = "CIN"
    sheet2.cell(row=1, column=4).value = "CNE"
    sheet2.cell(row=1, column=5).value = "Email"
    sheet2.cell(row=1, column=6).value = "Téléphone"
    sheet2.cell(row=1, column=7).value = "Classe"
    sheet2.cell(row=1, column=8).value = "Filière"

    sheet3.column_dimensions['A'].width = 30
    sheet3.column_dimensions['B'].width = 10
    sheet3.column_dimensions['C'].width = 10
    sheet3.column_dimensions['D'].width = 20
    sheet3.column_dimensions['E'].width = 20
    sheet3.column_dimensions['F'].width = 40
    sheet3.column_dimensions['G'].width = 50

    sheet3.cell(row=1, column=1).value = "Nom"
    sheet3.cell(row=1, column=2).value = "Prénom"
    sheet3.cell(row=1, column=3).value = "CIN"
    sheet3.cell(row=1, column=4).value = "CNE"
    sheet3.cell(row=1, column=5).value = "Email"
    sheet3.cell(row=1, column=6).value = "Téléphone"
    sheet3.cell(row=1, column=7).value = "Classe"
    sheet3.cell(row=1, column=8).value = "Filière"
  
  
def focus1(event): 
    prenom_field.focus_set() 
  
  
def focus2(event): 
    cin_field.focus_set() 
  
  
def focus3(event):  
    cne_field.focus_set() 
  
  
def focus4(event): 
    telephone_field.focus_set() 
  
  
def focus5(event): 
    email_field.focus_set() 
  

def focus6(event): 
    classe_field.focus_set() 
    
def focus7(event): 
    filiere_field.focus_set() 
  
def onselectclasse(evt):
    
    filiere_field.delete(0,END)
    w = evt.widget
    index = int(w.curselection()[0])
    filieres = []
    if index == 0 or index == 1 :
        filieres = ["GI","TM","TMIQ","GIM"]
    else :
        filieres = ["Mecatronique","MI","GCF","MQSE"]
    for item in filieres : filiere_field.insert(END, item)
  
def clear(): 
      
    nom_field.delete(0, END) 
    prenom_field.delete(0, END) 
    cin_field.delete(0, END) 
    cne_field.delete(0, END) 
    telephone_field.delete(0, END) 
    email_field.delete(0, END) 
  
def insert(): 
      
    if (nom_field.get() == "" and
        prenom_field.get() == "" and
        cin_field.get() == "" and
        cne_field.get() == "" and
        telephone_field.get() == "" and
        email_field.get() == "" and
        classe_field.get(classe_field.curselection()[0]) == "" and
        filiere_field.get(filiere_field.curselection()[0]) == "") : 
              
        print("empty input") 
  
    else: 
        if (classe_field.curselection()[0] == 0) :
            current_row = sheet.max_row
            current_column = sheet.max_column


            sheet.cell(row=current_row + 1, column=1).value = nom_field.get()
            sheet.cell(row=current_row + 1, column=2).value = prenom_field.get()
            sheet.cell(row=current_row + 1, column=3).value = cin_field.get()
            sheet.cell(row=current_row + 1, column=4).value = cne_field.get()
            sheet.cell(row=current_row + 1, column=5).value = email_field.get()
            sheet.cell(row=current_row + 1, column=6).value = telephone_field.get()
            sheet.cell(row=current_row + 1, column=7).value = classe_field.get(classe_field.curselection()[0])
            sheet.cell(row=current_row + 1, column=8).value = filiere_field.get(filiere_field.curselection()[0])

            wb.save('1ER.xlsx')

            nom_field.focus_set()

            clear()
        if (classe_field.curselection()[0] == 1):
            current_row = sheet2.max_row
            current_column = sheet2.max_column

            sheet2.cell(row=current_row + 1, column=1).value = nom_field.get()
            sheet2.cell(row=current_row + 1, column=2).value = prenom_field.get()
            sheet2.cell(row=current_row + 1, column=3).value = cin_field.get()
            sheet2.cell(row=current_row + 1, column=4).value = cne_field.get()
            sheet2.cell(row=current_row + 1, column=5).value = email_field.get()
            sheet2.cell(row=current_row + 1, column=6).value = telephone_field.get()
            sheet2.cell(row=current_row + 1, column=7).value = classe_field.get(classe_field.curselection()[0])
            sheet2.cell(row=current_row + 1, column=8).value = filiere_field.get(filiere_field.curselection()[0])

            wb2.save('2EME.xlsx')

            nom_field.focus_set()

            clear()
        if (classe_field.curselection()[0] == 2):
            current_row = sheet3.max_row
            current_column = sheet3.max_column

            sheet3.cell(row=current_row + 1, column=1).value = nom_field.get()
            sheet3.cell(row=current_row + 1, column=2).value = prenom_field.get()
            sheet3.cell(row=current_row + 1, column=3).value = cin_field.get()
            sheet3.cell(row=current_row + 1, column=4).value = cne_field.get()
            sheet3.cell(row=current_row + 1, column=5).value = email_field.get()
            sheet3.cell(row=current_row + 1, column=6).value = telephone_field.get()
            sheet3.cell(row=current_row + 1, column=7).value = classe_field.get(classe_field.curselection()[0])
            sheet3.cell(row=current_row + 1, column=8).value = filiere_field.get(filiere_field.curselection()[0])

            wb3.save('LP.xlsx')

            nom_field.focus_set()

            clear()


if __name__ == "__main__": 
      
    root = Tk() 
  
    root.configure(background='light yellow') 
  
    root.title("Ecole Supérieur de Technologie Safi") 
  
    root.geometry("500x500") 
  
    excel() 
  
    heading = Label(root, text="Ecole Supérieur de Technologie Safi", bg="light yellow") 
 
    nom = Label(root, text="Nom", bg="light yellow") 
  
    prenom = Label(root, text="Prénom", bg="light yellow") 
  
    cin = Label(root, text="CIN", bg="light yellow") 
  
    cne = Label(root, text="CNE", bg="light yellow") 
  
    email = Label(root, text="Email", bg="light yellow") 
  
    tele = Label(root, text="Téléphone", bg="light yellow") 
    
    classe = Label(root, text="Classe", bg="light yellow") 
    
    filiere = Label(root, text="Filiere", bg="light yellow") 
  
  
    heading.grid(row=0, column=1) 
    nom.grid(row=1, column=0) 
    prenom.grid(row=2, column=0) 
    cin.grid(row=3, column=0) 
    cne.grid(row=4, column=0) 
    tele.grid(row=5, column=0) 
    email.grid(row=6, column=0) 
    classe.grid(row=7, column=0) 
    filiere.grid(row=8, column=0) 

    
    nom_field = Entry(root) 
    prenom_field = Entry(root) 
    cin_field = Entry(root) 
    cne_field = Entry(root) 
    telephone_field = Entry(root) 
    email_field = Entry(root)  
    classe_field = Listbox(root,selectmode=SINGLE,exportselection=False)
    filiere_field = Listbox(root,selectmode=SINGLE,exportselection=False)

    for item in ["1 ère année", "2 ème année", "LP"]:
        classe_field.insert(END, item)

    classe_field.bind("<<ListboxSelect>>",onselectclasse)

    nom_field.bind("<Return>", focus1) 
  
    prenom_field.bind("<Return>", focus2) 
  
    cin_field.bind("<Return>", focus3) 
  
    cne_field.bind("<Return>", focus4) 
   
    telephone_field.bind("<Return>", focus5) 
  
    email_field.bind("<Return>", focus6)
    
    classe_field.bind("<Return>", focus7)
  
    nom_field.grid(row=1, column=1, ipadx="100") 
    prenom_field.grid(row=2, column=1, ipadx="100") 
    cin_field.grid(row=3, column=1, ipadx="100") 
    cne_field.grid(row=4, column=1, ipadx="100") 
    telephone_field.grid(row=5, column=1, ipadx="100") 
    email_field.grid(row=6, column=1, ipadx="100") 
    classe_field.grid(row=7, column=1, ipadx="100") 
    filiere_field.grid(row=8, column=1, ipadx="100") 

  
    excel() 
  
    submit = Button(root, text="Submit", fg="Black", 
                            bg="light green", command=insert) 
    submit.grid(row=9, column=1) 
  
    root.mainloop() 